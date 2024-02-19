"""
更新日期：
-.-.-

功能：
遍历一个文件夹，获取所有文件、文件夹的大小，并导出一个Excel表
"""

import os
import natsort
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill
import time


class DataDict(dict):
    """自定义字典，存放路径对应的属性（类型、大小、修改日期等）"""

    def __init__(self, the_type='file', *args, **kwargs):
        super().__init__(*args, **kwargs)

        # 设置默认键值对
        self['type'] = the_type  # 类型，file/folder
        self['size'] = 0  # 大小，文件为其大小，文件夹为其内部所有文件总大小
        self['name'] = ''  # 文件名

    def add_size(self, size):
        """添加大小属性"""
        if self['type'] == 'file':
            self['size'] = size
        elif self['type'] == 'folder':
            self['size'] += size


def get_parent_dirs(input_path):
    """获取一个路径的所有上级目录路径"""
    parent_dirs = []

    while True:
        parent_dirpath, filename = os.path.split(input_path)
        if filename:
            parent_dirs.append(parent_dirpath)
        else:
            break

        input_path = parent_dirpath

    # 反转列表顺序，使得最上级目录在最前面
    parent_dirs = parent_dirs[::-1]

    return parent_dirs


def get_dir_structure(dirpath: str):
    """获取一个文件夹路径的内部文件结构（含文件夹，且置顶）"""
    dir_list = []
    file_list = []
    for filename in natsort.natsorted(os.listdir(dirpath), alg=natsort.PATH | natsort.IGNORECASE):
        fullpath = os.path.normpath(os.path.join(dirpath, filename))
        if os.path.isdir(fullpath):
            dir_list.append(fullpath)
            dir_list += get_dir_structure(fullpath)  # 递归
        else:
            file_list.append(fullpath)

    # 整合
    join_list = dir_list + file_list

    return join_list


def walk_path(paths):
    """遍历文件路径"""
    # 检查传参格式，str转list
    if type(paths) is str:
        paths = [paths]
    # 剔除非文件夹路径，防止os.walk报错
    paths_checked = [i for i in paths if os.path.isdir(i)]

    # 获取内部结构（含文件夹）
    dir_structure = []
    for dirpath in natsort.natsorted(paths_checked):
        dir_structure += get_dir_structure(dirpath)

    return dir_structure


def calc_filesize(path_list):
    """计算文件大小并汇总"""
    # 计算文件大小，转为字典（key文件路径，value文件大小）
    # 并建立处理的路径的每层上级目录的字典（key文件夹路径，value文件夹内部文件总大小），进行文件夹大小累加
    path_data_dict = {}  # 文件属性字典，key为路径，value为对应的属性字典
    for path in path_list:
        if os.path.isfile(path):
            size = os.path.getsize(path)
            filename = os.path.split(path)[1]
            data_file = DataDict()
            data_file.add_size(size)
            data_file['name'] = filename
            path_data_dict[path] = data_file

            # 处理上级目录，大小累加入字典
            parent_dirs = get_parent_dirs(path)
            for p_dir in parent_dirs:
                if p_dir not in path_data_dict:
                    dirname = os.path.split(p_dir)[1]
                    data_dir = DataDict('folder')
                    data_dir['name'] = dirname
                    path_data_dict[p_dir] = data_dir
                path_data_dict[p_dir].add_size(size)
        else:
            if path not in path_data_dict:
                dirname = os.path.split(path)[1]
                data_dir2 = DataDict('folder')
                data_dir2['name'] = dirname
                path_data_dict[path] = data_dir2

    return path_data_dict


def join_structure_size(structure_list, path_data_dict):
    """将文件夹结构list和属性dict合并，生成最终结果的字典"""
    # key为路径，value为对应的属性字典
    structure_data_dict = {key: path_data_dict[key] for key in structure_list}

    return structure_data_dict


def save_to_excel(data_dict):
    """保存结果到excel"""
    localtime = time.strftime('%Y-%m-%d %H_%M_%S', time.localtime())
    result_file = f'遍历结果 {localtime}.xlsx'
    Workbook().save(result_file)

    wb = load_workbook(result_file)
    ws = wb.active

    ws['A1'] = '路径'
    ws['B1'] = '类型'
    ws['C1'] = '文件名'
    ws['D1'] = '大小/MB'
    ws.column_dimensions['C'].width = 40

    row = 1
    for path, data in data_dict.items():
        row += 1
        ws[f'A{row}'] = path
        ws[f'A{row}'].hyperlink = path
        ws[f'A{row}'].font = styles.Font(underline="single", color="0000FF")
        ws[f'B{row}'] = data['type']
        ws[f'C{row}'] = data['name']
        ws[f'D{row}'] = round(data['size'] / 1024 / 1024, 2)

        if data['type'] == 'folder':
            color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws[f'A{row}'].fill = color
            ws[f'B{row}'].fill = color
            ws[f'C{row}'].fill = color
            ws[f'D{row}'].fill = color

    wb.save(result_file)

    return result_file


def main():
    while True:
        info = """###文件夹遍历###\n###输入文件夹路径后遍历文件并生成excel自动打开###\n----------------\n"""
        print(info)

        path = input('输入路径并回车：')

        if os.path.isdir(path):
            dir_structure = walk_path(path)
            path_data_dict = calc_filesize(dir_structure)
            final_data_dict = join_structure_size(dir_structure, path_data_dict)
            result_file = save_to_excel(final_data_dict)
            os.startfile(result_file)
            print('已完成\n---------------\n')
        else:
            print('路径不正确，请重试')


if __name__ == '__main__':
    main()
